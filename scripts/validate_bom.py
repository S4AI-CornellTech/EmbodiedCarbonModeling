#!/usr/bin/env python3
"""
validate_bom.py

Small helper to validate ACT BOM YAML files and (optionally) compare them.

Usage:
  python scripts/validate_bom.py act/boms/RPi_Pico_W_Official.yaml
  python scripts/validate_bom.py act/boms/RPi_Pico_W_Official.yaml --compare act/boms/RPi_Pico_W_Officialv1.yaml
  python scripts/validate_bom.py act/boms/RPi_Pico_W_Official.yaml --compare act/boms/RPi_Pico_W_Officialv1.yaml --exact-text
  python scripts/validate_bom.py act/boms/RPi_Pico_W_Official.yaml --excel data/raw/RPi_Pico_W_Official.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any


REQUIRED_TOP_LEVEL_KEYS = {"name", "description"}
OPTIONAL_TOP_LEVEL_KEYS = {"silicon", "materials", "passives", "imports", "owner"}


def _load_yaml(path: Path) -> Any:
    try:
        import yaml  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "PyYAML is required to parse/validate BOMs. Install it with `pip install PyYAML` "
            "or run with `--exact-text` (compare-only)."
        ) from e

    with path.open("r", encoding="utf-8") as handle:
        return yaml.load(handle, Loader=yaml.FullLoader)


def _is_mapping(x: Any) -> bool:
    return isinstance(x, dict)


def validate_bom_dict(bom: Any, *, filename: str = "<bom>") -> list[str]:
    errors: list[str] = []

    if not _is_mapping(bom):
        return [f"{filename}: top-level YAML must be a mapping/dict"]

    missing = REQUIRED_TOP_LEVEL_KEYS - set(bom.keys())
    if missing:
        errors.append(f"{filename}: missing required top-level keys: {sorted(missing)}")

    unknown = set(bom.keys()) - (REQUIRED_TOP_LEVEL_KEYS | OPTIONAL_TOP_LEVEL_KEYS)
    if unknown:
        errors.append(f"{filename}: unknown top-level keys: {sorted(unknown)}")

    for section in ("silicon", "materials", "passives", "imports"):
        if section in bom and bom[section] is not None and not _is_mapping(bom[section]):
            errors.append(f"{filename}: `{section}` must be a mapping/dict if present")

    for section in ("materials", "passives"):
        sec = bom.get(section) or {}
        if not _is_mapping(sec):
            continue
        for comp_name, comp in sec.items():
            if not _is_mapping(comp):
                errors.append(f"{filename}: `{section}.{comp_name}` must be a mapping/dict")
                continue
            if "category" not in comp:
                errors.append(f"{filename}: `{section}.{comp_name}` missing `category`")
            if "quantity" in comp and comp["quantity"] is not None and not isinstance(
                comp["quantity"], int
            ):
                errors.append(
                    f"{filename}: `{section}.{comp_name}.quantity` must be an int if present"
                )

    silicon = bom.get("silicon") or {}
    if _is_mapping(silicon):
        for chip_name, chip in silicon.items():
            if not _is_mapping(chip):
                errors.append(f"{filename}: `silicon.{chip_name}` must be a mapping/dict")
                continue
            if "n_ics" in chip and chip["n_ics"] is not None and not isinstance(
                chip["n_ics"], int
            ):
                errors.append(f"{filename}: `silicon.{chip_name}.n_ics` must be an int")

    return errors


def _deep_sort(x: Any) -> Any:
    if isinstance(x, dict):
        return {k: _deep_sort(x[k]) for k in sorted(x.keys())}
    if isinstance(x, list):
        return [_deep_sort(v) for v in x]
    return x


def _norm_header(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()


def _as_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def _split_tags(tag_cell: Any) -> list[str]:
    tag_s = _as_str(tag_cell)
    if not tag_s:
        return []
    return [t.strip() for t in tag_s.split(",") if t.strip()]


def _key_suffix(value: str) -> str:
    return value.strip().replace(" ", "").replace("/", "_").replace("\\", "_")


def _load_excel_entries(excel_file: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required to validate against Excel. Install with `pip install openpyxl`."
        ) from e

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active

    header_row = None
    header_map: dict[str, int] = {}
    for r in range(1, min(15, ws.max_row) + 1):
        row = [_norm_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "quantity" in row and any(h.startswith("value") for h in row):
            header_row = r
            for idx, h in enumerate(row, start=1):
                if h in ("tag", "quantity", "component category", "component description"):
                    header_map[h] = idx
                elif h.startswith("value"):
                    header_map["value"] = idx
            break

    if header_row is None or "quantity" not in header_map or "value" not in header_map:
        raise ValueError(f"Could not find a header row with Quantity/Value in {excel_file}")

    out: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        tag_cell = ws.cell(r, header_map.get("tag", 1)).value if "tag" in header_map else None
        qty_cell = ws.cell(r, header_map["quantity"]).value
        value_cell = ws.cell(r, header_map["value"]).value

        if tag_cell is None and qty_cell is None and value_cell is None:
            continue

        out.append(
            {
                "row": r,
                "tags": _split_tags(tag_cell),
                "tag_raw": tag_cell,
                "quantity": int(qty_cell) if isinstance(qty_cell, (int, float)) and qty_cell is not None else None,
                "value": _as_str(value_cell),
            }
        )
    return out


def _expected_key_for_pico_entry(entry: dict[str, Any]) -> tuple[str | None, str | None]:
    # Returns (section, key) or (None, None) if entry doesn't map.
    tags: list[str] = entry["tags"]
    if not tags:
        return (None, None)

    base_tag = tags[0]
    value = entry["value"]
    suffix = _key_suffix(value) if value else ""

    if _as_str(entry["tag_raw"]) == "1" and value.upper() == "PCB":
        return (None, None)

    if base_tag.startswith("U"):
        return ("silicon", f"{base_tag}.{suffix}" if suffix else base_tag)

    if base_tag == "SH1":
        return ("materials", f"SH1.{suffix}" if suffix else "SH1")

    if base_tag.startswith("TP"):
        return ("materials", f"connector.TP1.{suffix}" if suffix else "connector.TP1")

    # Passives
    if base_tag.startswith("C"):
        return ("passives", f"capacitor.{base_tag}.{suffix}" if suffix else f"capacitor.{base_tag}")
    if base_tag.startswith("R"):
        return ("passives", f"resistor.{base_tag}.{suffix}" if suffix else f"resistor.{base_tag}")
    if base_tag.startswith("D"):
        return ("passives", f"diode.{base_tag}.{suffix}" if suffix else f"diode.{base_tag}")
    if base_tag.startswith("L"):
        return ("passives", f"inductor.{base_tag}.{suffix}" if suffix else f"inductor.{base_tag}")
    if base_tag.startswith("J"):
        return ("passives", f"connector.{base_tag}.{suffix}" if suffix else f"connector.{base_tag}")
    if base_tag.startswith("SW"):
        return ("passives", f"switch.{base_tag}.{suffix}" if suffix else f"switch.{base_tag}")
    if base_tag.startswith(("Q", "X")):
        return ("passives", f"active.{base_tag}.{suffix}" if suffix else f"active.{base_tag}")
    if base_tag.startswith("FB"):
        return ("passives", f"other.{base_tag}.{suffix}" if suffix else f"other.{base_tag}")
    if base_tag.startswith("FL"):
        return ("passives", f"active.{base_tag}.{suffix}" if suffix else f"active.{base_tag}")

    return (None, None)


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate an ACT BOM YAML file.")
    parser.add_argument("bom", type=Path, help="Path to BOM YAML file.")
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="Optional: validate that the BOM contains entries from this Excel file.",
    )
    parser.add_argument(
        "--compare",
        type=Path,
        default=None,
        help="Optional: compare BOM content against another BOM YAML file.",
    )
    parser.add_argument(
        "--exact-text",
        action="store_true",
        help="If set with --compare, compare file bytes (not parsed YAML).",
    )
    args = parser.parse_args()

    if not args.bom.exists():
        print(f"Error: BOM file not found: {args.bom}", file=sys.stderr)
        return 2

    bom_data = None
    if not (args.exact_text and args.compare is not None):
        try:
            bom_data = _load_yaml(args.bom)
        except RuntimeError as e:
            print(f"Error: {e}", file=sys.stderr)
            return 2

        errors = validate_bom_dict(bom_data, filename=str(args.bom))
        if errors:
            print("Validation failed:", file=sys.stderr)
            for e in errors:
                print(f"- {e}", file=sys.stderr)
            return 1

    print(f"OK: Valid BOM: {args.bom}")

    if args.excel is not None:
        if not args.excel.exists():
            print(f"Error: Excel file not found: {args.excel}", file=sys.stderr)
            return 2

        try:
            excel_entries = _load_excel_entries(args.excel)
        except Exception as e:
            print(f"Error reading Excel: {e}", file=sys.stderr)
            return 2

        # Only enforce tag-based matching (RPi4 sheet has no tags today).
        tagged = [e for e in excel_entries if e["tags"]]
        if tagged:
            stem = args.excel.stem
            if stem == "RPi_Pico_W_Official":
                for e in tagged:
                    section, key = _expected_key_for_pico_entry(e)
                    if section is None or key is None:
                        continue
                    section_dict = (bom_data or {}).get(section) or {}
                    if key not in section_dict:
                        print(
                            f"ERROR: Missing Excel entry in BOM (row {e['row']}): {section}.{key}",
                            file=sys.stderr,
                        )
                        return 1
                    expected_qty = e["quantity"]
                    if expected_qty is not None:
                        if section == "silicon":
                            actual_qty = (section_dict[key] or {}).get("n_ics")
                        else:
                            actual_qty = (section_dict[key] or {}).get("quantity")
                        if actual_qty != expected_qty:
                            print(
                                f"ERROR: Quantity mismatch for {section}.{key}: expected {expected_qty}, got {actual_qty}",
                                file=sys.stderr,
                            )
                            return 1
            else:
                print(
                    f"Warning: Excel validation for {stem} is tag-less or not implemented; skipping strict match."
                )
        else:
            print("Warning: Excel has no Tag column values; skipping strict BOM match.")

    if args.compare is not None:
        if not args.compare.exists():
            print(f"Error: compare file not found: {args.compare}", file=sys.stderr)
            return 2

        if args.exact_text:
            same = args.bom.read_bytes() == args.compare.read_bytes()
        else:
            try:
                other_data = _load_yaml(args.compare)
            except RuntimeError as e:
                print(f"Error: {e}", file=sys.stderr)
                return 2
            other_errors = validate_bom_dict(other_data, filename=str(args.compare))
            if other_errors:
                print("Compare target failed validation:", file=sys.stderr)
                for e in other_errors:
                    print(f"- {e}", file=sys.stderr)
                return 1

            same = _deep_sort(bom_data) == _deep_sort(other_data)

        if not same:
            print(f"ERROR: BOMs differ: {args.bom} vs {args.compare}", file=sys.stderr)
            return 1

        print(f"OK: BOMs match: {args.bom} == {args.compare}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
